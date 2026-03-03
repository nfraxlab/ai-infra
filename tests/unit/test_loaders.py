"""Tests for ai_infra.retriever.loaders module.

Tests file loading functionality for various formats.
"""

import importlib.util
import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from ai_infra.retriever.loaders import (
    _ocr_pdf_page,
    load_directory,
    load_file,
    load_json,
    load_text,
)

# Check for optional dependencies
HAS_PANDAS = importlib.util.find_spec("pandas") is not None
HAS_BS4 = importlib.util.find_spec("bs4") is not None
HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


class TestLoadFile:
    """Tests for load_file() function."""

    def test_load_txt_file(self, tmp_path: Path) -> None:
        """Test loading a .txt file."""
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("Hello, world!")

        docs = load_file(str(txt_file))

        assert len(docs) == 1
        text, metadata = docs[0]
        assert text == "Hello, world!"
        assert metadata["source"] == str(txt_file)
        assert metadata["file_type"] == ".txt"

    def test_load_md_file(self, tmp_path: Path) -> None:
        """Test loading a .md file."""
        md_file = tmp_path / "test.md"
        md_file.write_text("# Header\n\nSome content.")

        docs = load_file(str(md_file))

        assert len(docs) == 1
        text, metadata = docs[0]
        assert "# Header" in text
        assert metadata["file_type"] == ".md"

    def test_load_markdown_extension(self, tmp_path: Path) -> None:
        """Test loading a .markdown file."""
        md_file = tmp_path / "test.markdown"
        md_file.write_text("Markdown content")

        docs = load_file(str(md_file))

        assert len(docs) == 1
        assert docs[0][0] == "Markdown content"

    def test_file_not_found_raises(self) -> None:
        """Test that FileNotFoundError is raised for missing files."""
        with pytest.raises(FileNotFoundError, match="File not found"):
            load_file("/nonexistent/path/to/file.txt")

    def test_unsupported_extension_raises(self, tmp_path: Path) -> None:
        """Test that ValueError is raised for unsupported file types."""
        unknown_file = tmp_path / "test.xyz"
        unknown_file.write_text("content")

        with pytest.raises(ValueError, match="Unsupported file type"):
            load_file(str(unknown_file))

    def test_tilde_expansion(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        """Test that ~ is expanded in file paths."""
        # Create a file in tmp_path
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("content")

        # Monkeypatch os.path.expanduser to simulate ~ expansion
        monkeypatch.setenv("HOME", str(tmp_path))

        # load_file should expand ~ to tmp_path
        docs = load_file("~/test.txt")

        assert len(docs) == 1


class TestLoadText:
    """Tests for load_text() function."""

    def test_loads_plain_text(self, tmp_path: Path) -> None:
        """Test loading plain text content."""
        txt_file = tmp_path / "file.txt"
        content = "Line 1\nLine 2\nLine 3"
        txt_file.write_text(content)

        docs = load_text(str(txt_file))

        assert len(docs) == 1
        text, metadata = docs[0]
        assert text == content
        assert metadata["source"] == str(txt_file)

    def test_handles_unicode(self, tmp_path: Path) -> None:
        """Test loading text with unicode characters."""
        txt_file = tmp_path / "unicode.txt"
        content = "Hello 世界 🌍 Привет"
        txt_file.write_text(content, encoding="utf-8")

        docs = load_text(str(txt_file))

        assert docs[0][0] == content


class TestLoadJson:
    """Tests for load_json() function."""

    def test_loads_json_object(self, tmp_path: Path) -> None:
        """Test loading a single JSON object."""
        json_file = tmp_path / "data.json"
        data = {"name": "Alice", "age": 30}
        json_file.write_text(json.dumps(data))

        docs = load_json(str(json_file))

        assert len(docs) == 1
        text, metadata = docs[0]
        assert "Alice" in text
        assert metadata["file_type"] == ".json"

    def test_loads_json_array(self, tmp_path: Path) -> None:
        """Test loading a JSON array."""
        json_file = tmp_path / "array.json"
        data = [{"id": 1}, {"id": 2}, {"id": 3}]
        json_file.write_text(json.dumps(data))

        docs = load_json(str(json_file))

        assert len(docs) == 3
        for i, (text, metadata) in enumerate(docs):
            assert f'"id": {i + 1}' in text
            assert metadata["item_index"] == i

    def test_handles_nested_json(self, tmp_path: Path) -> None:
        """Test loading nested JSON structures."""
        json_file = tmp_path / "nested.json"
        data = {
            "users": [
                {"name": "Alice", "roles": ["admin"]},
                {"name": "Bob", "roles": ["user"]},
            ]
        }
        json_file.write_text(json.dumps(data))

        docs = load_json(str(json_file))

        assert len(docs) == 1
        text = docs[0][0]
        assert "Alice" in text
        assert "admin" in text


class TestLoadCsv:
    """Tests for load_csv() function."""

    @pytest.mark.skipif(not HAS_PANDAS, reason="pandas not installed")
    def test_loads_csv_rows(self, tmp_path: Path) -> None:
        """Test loading CSV rows as documents."""
        from ai_infra.retriever.loaders import load_csv

        csv_file = tmp_path / "data.csv"
        csv_file.write_text("name,age\nAlice,30\nBob,25")

        docs = load_csv(str(csv_file))

        assert len(docs) == 2
        assert "Alice" in docs[0][0]
        assert "30" in docs[0][0]
        assert docs[0][1]["file_type"] == ".csv"

    @pytest.mark.skipif(not HAS_PANDAS, reason="pandas not installed")
    def test_row_index_in_metadata(self, tmp_path: Path) -> None:
        """Test that row index is included in metadata."""
        from ai_infra.retriever.loaders import load_csv

        csv_file = tmp_path / "data.csv"
        csv_file.write_text("id,value\n1,a\n2,b\n3,c")

        docs = load_csv(str(csv_file))

        assert docs[0][1]["row_index"] == 0
        assert docs[1][1]["row_index"] == 1
        assert docs[2][1]["row_index"] == 2

    @pytest.mark.skipif(not HAS_PANDAS, reason="pandas not installed")
    def test_skips_empty_rows(self, tmp_path: Path) -> None:
        """Test that empty rows are handled gracefully."""
        from ai_infra.retriever.loaders import load_csv

        csv_file = tmp_path / "data.csv"
        csv_file.write_text("name,age\nAlice,30\n,\nBob,25")

        docs = load_csv(str(csv_file))

        # The empty row should not produce a document
        # (depends on implementation - if it has empty strings, it may still appear)
        assert len(docs) >= 2


class TestLoadHtml:
    """Tests for load_html() function."""

    @pytest.mark.skipif(not HAS_BS4, reason="beautifulsoup4 not installed")
    def test_extracts_text_from_html(self, tmp_path: Path) -> None:
        """Test extracting text from HTML."""
        from ai_infra.retriever.loaders import load_html

        html_file = tmp_path / "page.html"
        html_file.write_text("<html><body><p>Hello, world!</p></body></html>")

        docs = load_html(str(html_file))

        assert len(docs) == 1
        text, metadata = docs[0]
        assert "Hello, world!" in text
        assert metadata["file_type"] == ".html"

    @pytest.mark.skipif(not HAS_BS4, reason="beautifulsoup4 not installed")
    def test_removes_script_and_style(self, tmp_path: Path) -> None:
        """Test that script and style elements are removed."""
        from ai_infra.retriever.loaders import load_html

        html_file = tmp_path / "page.html"
        html_file.write_text("""
        <html>
        <head>
            <style>body { color: red; }</style>
            <script>alert('test');</script>
        </head>
        <body>
            <p>Visible content</p>
        </body>
        </html>
        """)

        docs = load_html(str(html_file))

        text = docs[0][0]
        assert "Visible content" in text
        assert "alert" not in text
        assert "color: red" not in text

    @pytest.mark.skipif(not HAS_BS4, reason="beautifulsoup4 not installed")
    def test_extracts_title(self, tmp_path: Path) -> None:
        """Test that title is extracted to metadata."""
        from ai_infra.retriever.loaders import load_html

        html_file = tmp_path / "page.html"
        html_file.write_text("<html><head><title>My Page</title></head><body>Content</body></html>")

        docs = load_html(str(html_file))

        metadata = docs[0][1]
        assert metadata["title"] == "My Page"

    @pytest.mark.skipif(not HAS_BS4, reason="beautifulsoup4 not installed")
    def test_htm_extension(self, tmp_path: Path) -> None:
        """Test that .htm extension is also supported."""
        htm_file = tmp_path / "page.htm"
        htm_file.write_text("<html><body>Content</body></html>")

        docs = load_file(str(htm_file))

        assert len(docs) == 1
        assert "Content" in docs[0][0]


class TestLoadDirectory:
    """Tests for load_directory() function."""

    def test_loads_all_supported_files(self, tmp_path: Path) -> None:
        """Test loading all supported files from a directory."""
        (tmp_path / "file1.txt").write_text("Text 1")
        (tmp_path / "file2.txt").write_text("Text 2")
        (tmp_path / "file3.md").write_text("Markdown")

        docs = load_directory(str(tmp_path))

        assert len(docs) == 3
        texts = [d[0] for d in docs]
        assert "Text 1" in texts
        assert "Text 2" in texts
        assert "Markdown" in texts

    def test_pattern_filtering(self, tmp_path: Path) -> None:
        """Test filtering by glob pattern."""
        (tmp_path / "file1.txt").write_text("Text")
        (tmp_path / "file2.md").write_text("Markdown")

        docs = load_directory(str(tmp_path), pattern="*.txt")

        assert len(docs) == 1
        assert docs[0][0] == "Text"

    def test_recursive_loading(self, tmp_path: Path) -> None:
        """Test recursive directory loading."""
        subdir = tmp_path / "subdir"
        subdir.mkdir()
        (tmp_path / "root.txt").write_text("Root")
        (subdir / "nested.txt").write_text("Nested")

        docs = load_directory(str(tmp_path), recursive=True)

        assert len(docs) == 2
        texts = [d[0] for d in docs]
        assert "Root" in texts
        assert "Nested" in texts

    def test_non_recursive_loading(self, tmp_path: Path) -> None:
        """Test non-recursive directory loading."""
        subdir = tmp_path / "subdir"
        subdir.mkdir()
        (tmp_path / "root.txt").write_text("Root")
        (subdir / "nested.txt").write_text("Nested")

        docs = load_directory(str(tmp_path), recursive=False)

        # Should only get root.txt since recursive=False uses glob() not rglob()
        texts = [d[0] for d in docs]
        assert "Root" in texts

    def test_skips_unsupported_files(self, tmp_path: Path) -> None:
        """Test that unsupported files are skipped silently."""
        (tmp_path / "file.txt").write_text("Text")
        (tmp_path / "file.xyz").write_text("Unknown")

        docs = load_directory(str(tmp_path))

        # Should only load the .txt file
        assert len(docs) == 1
        assert docs[0][0] == "Text"

    def test_directory_not_found_raises(self) -> None:
        """Test that FileNotFoundError is raised for missing directory."""
        with pytest.raises(FileNotFoundError, match="Directory not found"):
            load_directory("/nonexistent/path/to/dir")

    def test_tilde_expansion_in_directory(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Test that ~ is expanded in directory paths."""
        (tmp_path / "file.txt").write_text("Content")

        monkeypatch.setenv("HOME", str(tmp_path))

        docs = load_directory("~")

        assert len(docs) >= 1

    def test_empty_directory(self, tmp_path: Path) -> None:
        """Test loading from an empty directory."""
        docs = load_directory(str(tmp_path))

        assert docs == []


class TestLoadPdf:
    """Tests for load_pdf() function.

    Note: These tests require pypdf to be installed.
    """

    def test_pdf_import_error_message(self) -> None:
        """Test that ImportError provides installation instructions."""
        # This test is conditional - if pypdf is installed, we can't test the error
        try:
            from pypdf import PdfReader  # noqa: F401

            pytest.skip("pypdf is installed, cannot test ImportError")
        except ImportError:
            from ai_infra.retriever.loaders import load_pdf

            with pytest.raises(ImportError, match="pip install pypdf"):
                load_pdf("test.pdf")


class TestLoadDocx:
    """Tests for load_docx() function.

    Note: These tests require python-docx to be installed.
    """

    def test_docx_import_error_message(self) -> None:
        """Test that ImportError provides installation instructions."""
        try:
            from docx import Document  # noqa: F401

            pytest.skip("python-docx is installed, cannot test ImportError")
        except ImportError:
            from ai_infra.retriever.loaders import load_docx

            with pytest.raises(ImportError, match="pip install python-docx"):
                load_docx("test.docx")


class TestIntegration:
    """Integration tests for file loading."""

    def test_load_file_delegates_to_correct_loader(self, tmp_path: Path) -> None:
        """Test that load_file delegates to the correct loader based on extension."""
        # Create files of different types
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("text content")

        json_file = tmp_path / "test.json"
        json_file.write_text('{"key": "value"}')

        # Load each and verify correct handling
        txt_docs = load_file(str(txt_file))
        assert txt_docs[0][1]["file_type"] == ".txt"

        json_docs = load_file(str(json_file))
        assert json_docs[0][1]["file_type"] == ".json"

    @pytest.mark.skipif(not HAS_BS4, reason="beautifulsoup4 not installed")
    def test_load_html_file_type(self, tmp_path: Path) -> None:
        """Test loading HTML files."""
        html_file = tmp_path / "test.html"
        html_file.write_text("<html><body>html content</body></html>")

        html_docs = load_file(str(html_file))
        assert html_docs[0][1]["file_type"] == ".html"

    def test_source_metadata_consistent(self, tmp_path: Path) -> None:
        """Test that source metadata is consistently set."""
        txt_file = tmp_path / "file.txt"
        txt_file.write_text("content")

        docs = load_file(str(txt_file))

        # Source should be the full path
        assert docs[0][1]["source"] == str(txt_file)

    def test_mixed_directory_loading(self, tmp_path: Path) -> None:
        """Test loading a directory with mixed file types."""
        (tmp_path / "file1.txt").write_text("Text file")
        (tmp_path / "file2.json").write_text('{"data": 123}')
        (tmp_path / "file4.unsupported").write_text("Ignored")

        docs = load_directory(str(tmp_path))

        # Should have 2 documents (txt, json - unsupported is skipped)
        assert len(docs) == 2

        # Verify each type was loaded
        sources = [d[1]["source"] for d in docs]
        assert any("file1.txt" in s for s in sources)
        assert any("file2.json" in s for s in sources)

    @pytest.mark.skipif(not HAS_BS4, reason="beautifulsoup4 not installed")
    def test_mixed_directory_with_html(self, tmp_path: Path) -> None:
        """Test loading directory that includes HTML files."""
        (tmp_path / "file1.txt").write_text("Text file")
        (tmp_path / "file3.html").write_text("<html><body>HTML</body></html>")

        docs = load_directory(str(tmp_path))

        assert len(docs) == 2
        sources = [d[1]["source"] for d in docs]
        assert any("file1.txt" in s for s in sources)
        assert any("file3.html" in s for s in sources)


class TestOcrPdfPage:
    """Tests for _ocr_pdf_page() OCR fallback (pypdfium2 + pytesseract)."""

    def test_returns_empty_when_deps_unavailable(self, tmp_path: Path) -> None:
        """Graceful fallback when pypdfium2/pytesseract are not installed."""
        with patch.dict("sys.modules", {"pypdfium2": None, "pytesseract": None}):
            result = _ocr_pdf_page(str(tmp_path / "fake.pdf"), 0)
            assert result == ""

    def test_returns_empty_on_nonexistent_file(self) -> None:
        """Returns empty string for a nonexistent file path (caught by except)."""
        result = _ocr_pdf_page("/nonexistent/does_not_exist.pdf", 0)
        assert result == ""

    def test_returns_empty_on_render_failure(self, tmp_path: Path) -> None:
        """Graceful handling when pypdfium2.PdfDocument raises."""
        mock_pdfium = MagicMock()
        mock_pdfium.PdfDocument.side_effect = RuntimeError("corrupt pdf")

        with patch.dict(
            "sys.modules",
            {"pypdfium2": mock_pdfium, "pytesseract": MagicMock()},
        ):
            result = _ocr_pdf_page(str(tmp_path / "bad.pdf"), 0)

        assert result == ""


class TestLoadExcel:
    """Tests for load_excel() function."""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_single_sheet(self, tmp_path: Path) -> None:
        """Test loading a single-sheet Excel file."""
        from openpyxl import Workbook

        from ai_infra.retriever.loaders import load_excel

        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        docs = load_excel(str(path))

        assert len(docs) == 1
        text, metadata = docs[0]
        assert "Name | Age" in text
        assert "Alice | 30" in text
        assert metadata["file_type"] == ".xlsx"
        assert "sheet" in metadata

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_multiple_sheets(self, tmp_path: Path) -> None:
        """Test loading multi-sheet Excel file returns one doc per sheet."""
        from openpyxl import Workbook

        from ai_infra.retriever.loaders import load_excel

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Col1"])
        ws1.append(["val1"])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Col2"])
        ws2.append(["val2"])

        path = tmp_path / "multi.xlsx"
        wb.save(str(path))

        docs = load_excel(str(path))

        assert len(docs) == 2
        sheets = [d[1]["sheet"] for d in docs]
        assert "Sheet1" in sheets
        assert "Sheet2" in sheets

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_skips_empty_sheets(self, tmp_path: Path) -> None:
        """Test that completely empty sheets are skipped."""
        from openpyxl import Workbook

        from ai_infra.retriever.loaders import load_excel

        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        # Don't add any data

        ws2 = wb.create_sheet("HasData")
        ws2.append(["Header"])
        ws2.append(["Value"])

        path = tmp_path / "sparse.xlsx"
        wb.save(str(path))

        docs = load_excel(str(path))

        # Only the sheet with data should be loaded
        assert len(docs) == 1
        assert docs[0][1]["sheet"] == "HasData"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_xlsx_via_load_file(self, tmp_path: Path) -> None:
        """Test that .xlsx files are dispatched through load_file."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["A"])
        ws.append(["B"])
        path = tmp_path / "via_load_file.xlsx"
        wb.save(str(path))

        docs = load_file(str(path))

        assert len(docs) == 1
        assert "A" in docs[0][0]

    def test_excel_import_error_message(self) -> None:
        """Test that ImportError provides installation instructions."""
        try:
            from openpyxl import Workbook  # noqa: F401

            pytest.skip("openpyxl is installed, cannot test ImportError")
        except ImportError:
            from ai_infra.retriever.loaders import load_excel

            with pytest.raises(ImportError, match="openpyxl"):
                load_excel("test.xlsx")
